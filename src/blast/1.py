import openpyxl
n = str(6) 
xlsx = '/mnt/md0/Public/T3_T4/txsedb/T'+n+'SE.xlsx'
tX_positive = rf'/mnt/md0/Public/T3_T4/data/new_T{n}/pos'

wb = openpyxl.load_workbook(xlsx)
ws = wb.active
max_row = ws.max_row
n
raw_TX_training_fasta = tX_positive+'/T'+n+'_training.fasta'
bac_name = ['Escherichia coli','Pseudomonas','Burkholderia','Pseudomonas'][3]
raw_TX_Ralstonia_fasta = tX_positive+'/T'+n+'_'+bac_name+'.fasta'
raw_total_fasta =  tX_positive+'/T'+n+'_summary.fasta'

training_outfile = open(raw_TX_training_fasta,'w',encoding='utf-8')

val_outfile = open(raw_TX_Ralstonia_fasta, 'w',encoding='utf-8')
total_outfile = open(raw_total_fasta,'w',encoding='utf-8')
for  i in range(2,max_row+1):
    k = ws.cell(row = i ,column = 12).value
    protein_id = ws.cell(row=i, column=5).value
    sequence = ws.cell(row=i, column=14).value
    total_outfile.write(">"+protein_id+'\n'+sequence+'\n')
    if  bac_name in k :
        val_outfile.write(">"+protein_id+'\n'+sequence+'\n')
    elif  bac_name  not in k:
        training_outfile.write(">"+protein_id+'\n'+sequence+'\n')


val_outfile.close()
training_outfile.close()
total_outfile.close()
wb.save(xlsx)